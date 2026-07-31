import io
import uuid
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import NotFoundException
from app.schemas.skills.match import MatchReport, RiskLevel
from app.schemas.skills.parse import (
    ColumnVariant,
    ParseResult,
    SectionKind,
)
from app.services.match_service import MatchService
from app.services.parse_service import ParseService
from app.services.task_service import TaskService


_HEADER_FILL = PatternFill("solid", fgColor="18794E")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
_TITLE_FONT = Font(size=14, bold=True, color="1F2924")
_SECTION_FONT = Font(size=11, bold=True, color="18794E")
_BODY_ALIGN = Alignment(vertical="top", wrap_text=True)


class ExportService:
    @staticmethod
    async def build_xlsx(
        session: AsyncSession,
        task_id: uuid.UUID,
    ) -> tuple[str, bytes]:
        """生成投标分析报告工作簿，返回（文件名, xlsx 字节）。"""

        task = await TaskService.get(session, task_id)
        parse_record = await ParseService.get_latest_result(session, task_id)
        parse_result = ParseResult.model_validate(parse_record.result_json)

        match_report: MatchReport | None = None
        try:
            match_record = await MatchService.get_latest_result(session, task_id)
            if match_record.result_json is not None:
                match_report = MatchReport.model_validate(match_record.result_json)
        except NotFoundException:
            match_report = None

        workbook = Workbook()
        workbook.remove(workbook.active)
        ExportService._write_parse_sheets(workbook, parse_result)
        if match_report is not None:
            ExportService._write_match_sheets(workbook, match_report)

        project_name = parse_result.project_name or task.project_name or "未命名项目"
        filename = f"投标分析报告_{project_name}.xlsx"
        buffer = io.BytesIO()
        workbook.save(buffer)
        return filename, buffer.getvalue()

    @staticmethod
    def _write_parse_sheets(
        workbook: Workbook,
        parse_result: ParseResult,
    ) -> None:
        used_titles: set[str] = set()

        def sheet(title: str):
            safe = title[:28]
            base = safe
            index = 2
            while safe in used_titles:
                safe = f"{base}-{index}"
                index += 1
            used_titles.add(safe)
            return workbook.create_sheet(safe)

        for section in parse_result.template.sections:
            data = parse_result.data.get(section.id)
            if section.kind == SectionKind.GRID:
                rows: list[tuple[str, str]] = []
                if isinstance(data, dict):
                    for field in section.fields:
                        value = data.get(field.key)
                        rows.append((field.label, _format_value(value)))
                rows.append(("原始摘要", parse_result.raw_summary or "-"))
                rows.append(("模板版本", parse_result.template.version))
                ExportService._write_rows(
                    sheet(section.title),
                    ["字段", "内容"],
                    rows,
                )
            elif section.kind == SectionKind.TABLE:
                columns = _table_columns(section)
                headers = [label for _, label, _ in columns]
                rows = []
                for row in data if isinstance(data, list) else []:
                    rows.append(
                        [
                            _format_value(row.get(key))
                            for _, _, key in columns
                        ]
                    )
                ExportService._write_rows(sheet(section.title), headers, rows)
            elif section.kind == SectionKind.KEY_VALUE:
                rows = [
                    (str(key), _format_value(value))
                    for key, value in (data.items() if isinstance(data, dict) else [])
                ]
                ExportService._write_rows(
                    sheet(section.title),
                    ["维度", "内容"],
                    rows,
                )
            else:
                rows = [
                    (str(index), _format_value(item))
                    for index, item in enumerate(
                        data if isinstance(data, list) else [],
                        start=1,
                    )
                ]
                ExportService._write_rows(
                    sheet(section.title),
                    ["序号", "内容"],
                    rows,
                )

    @staticmethod
    def _write_match_sheets(
        workbook: Workbook,
        report: MatchReport,
    ) -> None:
        overview = workbook.create_sheet("资质匹配总览")
        ExportService._write_rows(
            overview,
            ["项目", "结果"],
            [
                (
                    "综合匹配分",
                    "-" if report.overall_match_score is None
                    else f"{report.overall_match_score:.2f}",
                ),
                ("匹配结论", report.summary),
            ],
        )

        groups = [
            ("已满足项", report.matched_items, "success"),
            ("缺失项", report.missing_items, "danger"),
            ("风险项", report.risk_items, "warning"),
        ]
        for title, items, _ in groups:
            ExportService._write_rows(
                workbook.create_sheet(title),
                ["类别", "招标要求", "公司现状", "风险等级", "说明"],
                [
                    [
                        item.category,
                        item.requirement,
                        item.company_status,
                        _risk_label(item.risk_level),
                        item.comment or "",
                    ]
                    for item in items
                ],
            )

        suggestions = workbook.create_sheet("处理建议")
        ExportService._write_rows(
            suggestions,
            ["序号", "建议"],
            [
                (str(index), suggestion)
                for index, suggestion in enumerate(report.suggestions, start=1)
            ],
        )

        checklist = workbook.create_sheet("材料准备清单")
        checklist_rows: list[list[str]] = []
        for item in report.missing_items:
            checklist_rows.append(
                [
                    item.category,
                    item.requirement,
                    "缺失",
                    "准备或补充对应证明材料",
                    item.comment or "",
                ]
            )
        for item in report.risk_items:
            checklist_rows.append(
                [
                    item.category,
                    item.requirement,
                    "风险",
                    "投标前核验原件与边界条件",
                    item.comment or "",
                ]
            )
        ExportService._write_rows(
            checklist,
            ["类别", "招标要求", "状态", "需要准备/核验", "说明"],
            checklist_rows,
        )

    @staticmethod
    def _write_rows(
        worksheet: Any,
        headers: list[str],
        rows: list[list[str] | tuple[str, ...]],
    ) -> None:
        worksheet.append(headers)
        for cell in worksheet[1]:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(vertical="center")
        for row in rows:
            worksheet.append(list(row))
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = _BODY_ALIGN
        widths = [len(header) for header in headers]
        for row in rows:
            for index, value in enumerate(row, start=1):
                if index <= len(widths):
                    widths[index - 1] = max(widths[index - 1], len(str(value)))
        for index, width in enumerate(widths, start=1):
            worksheet.column_dimensions[get_column_letter(index)].width = min(
                max(14, width + 4),
                60,
            )


def _table_columns(section: Any) -> list[tuple[str, str, str]]:
    columns: list[tuple[str, str, str]] = []
    for column in section.columns:
        columns.append((column.label, column.label, column.key))
        if column.variant == ColumnVariant.STACK and column.secondary_key:
            columns.append(
                (
                    column.secondary_prefix or column.secondary_key,
                    column.secondary_prefix or column.secondary_key,
                    column.secondary_key,
                )
            )
    return columns


def _format_value(value: Any) -> str:
    if value is None or value == "":
        return "-"
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (list, tuple)):
        return "；".join(_format_value(item) for item in value)
    if isinstance(value, dict):
        return "；".join(f"{key}：{_format_value(item)}" for key, item in value.items())
    return str(value)


def _risk_label(level: RiskLevel) -> str:
    return {
        RiskLevel.NONE: "无风险",
        RiskLevel.LOW: "低风险",
        RiskLevel.MEDIUM: "中风险",
        RiskLevel.HIGH: "高风险",
    }.get(level, str(level))
