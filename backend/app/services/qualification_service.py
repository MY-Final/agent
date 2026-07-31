import io
import uuid
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any, TypeVar

from fastapi import UploadFile
from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy import and_, func, not_, or_, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import (
    AppException,
    ConflictException,
    NotFoundException,
    StorageException,
)
from app.core.minio import MinIOStorage
from app.models.qualification import (
    CertificateStatus,
    CompanyProfile,
    PerformanceRecord,
    PersonnelCertificate,
    QualificationCertificate,
)
from app.schemas.qualification import (
    CompanyProfileCreate,
    CompanyProfileUpdate,
    ExpiryWarningItem,
    ExpiryWarningsRead,
    PerformanceRecordCreate,
    PerformanceRecordUpdate,
    PersonnelCertificateCreate,
    PersonnelCertificateUpdate,
    QualificationImportResult,
    QualificationCertificateCreate,
    QualificationCertificateUpdate,
)


ModelT = TypeVar("ModelT")


class QualificationService:
    @staticmethod
    async def create_certificate(
        session: AsyncSession,
        payload: QualificationCertificateCreate,
    ) -> QualificationCertificate:
        record = QualificationCertificate(**payload.model_dump())
        session.add(record)
        await session.commit()
        await session.refresh(record)
        return record

    @staticmethod
    async def list_certificates(
        session: AsyncSession,
        *,
        name: str | None,
        status: CertificateStatus | None,
        is_valid: bool | None,
        page: int,
        page_size: int,
    ) -> tuple[list[QualificationCertificate], int]:
        filters: list[Any] = []
        if name:
            filters.append(QualificationCertificate.name.ilike(f"%{name.strip()}%"))
        if status is not None:
            filters.append(QualificationCertificate.status == status)
        if is_valid is not None:
            valid_condition = QualificationService._certificate_valid_condition()
            filters.append(valid_condition if is_valid else not_(valid_condition))
        return await QualificationService._paginate(
            session,
            QualificationCertificate,
            filters,
            page,
            page_size,
            QualificationCertificate.created_at.desc(),
        )

    @staticmethod
    async def get_certificate(
        session: AsyncSession,
        record_id: uuid.UUID,
    ) -> QualificationCertificate:
        record = await session.get(QualificationCertificate, record_id)
        if record is None:
            raise NotFoundException("资质证书不存在")
        return record

    @staticmethod
    async def update_certificate(
        session: AsyncSession,
        record_id: uuid.UUID,
        payload: QualificationCertificateUpdate,
    ) -> QualificationCertificate:
        record = await QualificationService.get_certificate(session, record_id)
        values = payload.model_dump(exclude_unset=True)
        QualificationService._validate_merged_date_range(record, values, "有效期")
        QualificationService._apply_values(record, values)
        await session.commit()
        await session.refresh(record)
        return record

    @staticmethod
    async def delete_certificate(
        session: AsyncSession,
        storage: MinIOStorage,
        record_id: uuid.UUID,
    ) -> None:
        record = await QualificationService.get_certificate(session, record_id)
        await QualificationService._delete_with_attachment(session, storage, record)

    @staticmethod
    async def create_performance(
        session: AsyncSession,
        payload: PerformanceRecordCreate,
    ) -> PerformanceRecord:
        record = PerformanceRecord(**payload.model_dump())
        session.add(record)
        await session.commit()
        await session.refresh(record)
        return record

    @staticmethod
    async def list_performances(
        session: AsyncSession,
        *,
        keyword: str | None,
        min_amount: Decimal | None,
        max_amount: Decimal | None,
        is_completed: bool | None,
        page: int,
        page_size: int,
    ) -> tuple[list[PerformanceRecord], int]:
        if min_amount is not None and max_amount is not None and min_amount > max_amount:
            raise AppException(
                "最小业绩金额不能大于最大业绩金额",
                code=42222,
                status_code=422,
            )
        filters: list[Any] = []
        if keyword:
            pattern = f"%{keyword.strip()}%"
            filters.append(
                or_(
                    PerformanceRecord.project_name.ilike(pattern),
                    PerformanceRecord.owner_name.ilike(pattern),
                    PerformanceRecord.related_qualification.ilike(pattern),
                    PerformanceRecord.description.ilike(pattern),
                )
            )
        if min_amount is not None:
            filters.append(PerformanceRecord.project_amount >= min_amount)
        if max_amount is not None:
            filters.append(PerformanceRecord.project_amount <= max_amount)
        if is_completed is not None:
            filters.append(PerformanceRecord.is_completed == is_completed)
        return await QualificationService._paginate(
            session,
            PerformanceRecord,
            filters,
            page,
            page_size,
            PerformanceRecord.created_at.desc(),
        )

    @staticmethod
    async def get_performance(
        session: AsyncSession,
        record_id: uuid.UUID,
    ) -> PerformanceRecord:
        record = await session.get(PerformanceRecord, record_id)
        if record is None:
            raise NotFoundException("业绩记录不存在")
        return record

    @staticmethod
    async def update_performance(
        session: AsyncSession,
        record_id: uuid.UUID,
        payload: PerformanceRecordUpdate,
    ) -> PerformanceRecord:
        record = await QualificationService.get_performance(session, record_id)
        values = payload.model_dump(exclude_unset=True)
        QualificationService._validate_merged_date_range(
            record,
            values,
            "项目日期",
            start_field="start_date",
            end_field="end_date",
        )
        QualificationService._apply_values(record, values)
        await session.commit()
        await session.refresh(record)
        return record

    @staticmethod
    async def delete_performance(
        session: AsyncSession,
        storage: MinIOStorage,
        record_id: uuid.UUID,
    ) -> None:
        record = await QualificationService.get_performance(session, record_id)
        await QualificationService._delete_with_attachment(session, storage, record)

    @staticmethod
    async def create_personnel(
        session: AsyncSession,
        payload: PersonnelCertificateCreate,
    ) -> PersonnelCertificate:
        record = PersonnelCertificate(**payload.model_dump())
        session.add(record)
        await session.commit()
        await session.refresh(record)
        return record

    @staticmethod
    async def list_personnel(
        session: AsyncSession,
        *,
        person_name: str | None,
        cert_type: str | None,
        specialty: str | None,
        is_on_job: bool | None,
        is_valid: bool | None,
        page: int,
        page_size: int,
    ) -> tuple[list[PersonnelCertificate], int]:
        filters: list[Any] = []
        if person_name:
            filters.append(
                PersonnelCertificate.person_name.ilike(f"%{person_name.strip()}%")
            )
        if cert_type:
            filters.append(
                PersonnelCertificate.cert_type.ilike(f"%{cert_type.strip()}%")
            )
        if specialty:
            filters.append(
                PersonnelCertificate.specialty.ilike(f"%{specialty.strip()}%")
            )
        if is_on_job is not None:
            filters.append(PersonnelCertificate.is_on_job == is_on_job)
        if is_valid is not None:
            valid_condition = QualificationService._personnel_valid_condition()
            filters.append(valid_condition if is_valid else not_(valid_condition))
        return await QualificationService._paginate(
            session,
            PersonnelCertificate,
            filters,
            page,
            page_size,
            PersonnelCertificate.created_at.desc(),
        )

    @staticmethod
    async def get_personnel(
        session: AsyncSession,
        record_id: uuid.UUID,
    ) -> PersonnelCertificate:
        record = await session.get(PersonnelCertificate, record_id)
        if record is None:
            raise NotFoundException("人员证书不存在")
        return record

    @staticmethod
    async def update_personnel(
        session: AsyncSession,
        record_id: uuid.UUID,
        payload: PersonnelCertificateUpdate,
    ) -> PersonnelCertificate:
        record = await QualificationService.get_personnel(session, record_id)
        values = payload.model_dump(exclude_unset=True)
        QualificationService._validate_merged_date_range(record, values, "有效期")
        QualificationService._apply_values(record, values)
        await session.commit()
        await session.refresh(record)
        return record

    @staticmethod
    async def delete_personnel(
        session: AsyncSession,
        storage: MinIOStorage,
        record_id: uuid.UUID,
    ) -> None:
        record = await QualificationService.get_personnel(session, record_id)
        await QualificationService._delete_with_attachment(session, storage, record)

    @staticmethod
    async def create_company(
        session: AsyncSession,
        payload: CompanyProfileCreate,
    ) -> CompanyProfile:
        record = CompanyProfile(**payload.model_dump())
        session.add(record)
        try:
            await session.commit()
        except IntegrityError as exc:
            await session.rollback()
            raise ConflictException("公司名称已存在") from exc
        await session.refresh(record)
        return record

    @staticmethod
    async def list_companies(
        session: AsyncSession,
        *,
        company_name: str | None,
        page: int,
        page_size: int,
    ) -> tuple[list[CompanyProfile], int]:
        filters: list[Any] = []
        if company_name:
            filters.append(CompanyProfile.company_name.ilike(f"%{company_name.strip()}%"))
        return await QualificationService._paginate(
            session,
            CompanyProfile,
            filters,
            page,
            page_size,
            CompanyProfile.created_at.desc(),
        )

    @staticmethod
    async def get_company(
        session: AsyncSession,
        record_id: uuid.UUID,
    ) -> CompanyProfile:
        record = await session.get(CompanyProfile, record_id)
        if record is None:
            raise NotFoundException("公司信息不存在")
        return record

    @staticmethod
    async def update_company(
        session: AsyncSession,
        record_id: uuid.UUID,
        payload: CompanyProfileUpdate,
    ) -> CompanyProfile:
        record = await QualificationService.get_company(session, record_id)
        QualificationService._apply_values(record, payload.model_dump(exclude_unset=True))
        try:
            await session.commit()
        except IntegrityError as exc:
            await session.rollback()
            raise ConflictException("公司名称已存在") from exc
        await session.refresh(record)
        return record

    @staticmethod
    async def delete_company(session: AsyncSession, record_id: uuid.UUID) -> None:
        record = await QualificationService.get_company(session, record_id)
        await session.delete(record)
        await session.commit()

    @staticmethod
    async def get_knowledge_base(
        session: AsyncSession,
    ) -> tuple[
        list[QualificationCertificate],
        list[PerformanceRecord],
        list[PersonnelCertificate],
        list[CompanyProfile],
    ]:
        certificates = list(
            (await session.scalars(select(QualificationCertificate))).all()
        )
        performances = list((await session.scalars(select(PerformanceRecord))).all())
        personnel = list(
            (await session.scalars(select(PersonnelCertificate))).all()
        )
        companies = list((await session.scalars(select(CompanyProfile))).all())
        return certificates, performances, personnel, companies

    @staticmethod
    async def get_expiry_warnings(
        session: AsyncSession,
    ) -> ExpiryWarningsRead:
        """证书/人员证书的失效预警：临期、已过期、已撤销、离职。"""

        today = date.today()
        warning_cutoff = today + timedelta(days=90)
        certificates = list(
            (await session.scalars(select(QualificationCertificate))).all()
        )
        personnel = list(
            (await session.scalars(select(PersonnelCertificate))).all()
        )

        items: list[ExpiryWarningItem] = []
        for cert in certificates:
            days_left = (cert.valid_to - today).days if cert.valid_to else None
            if cert.status == CertificateStatus.REVOKED:
                status = "revoked"
            elif days_left is not None and days_left < 0:
                status = "expired"
            elif cert.status == CertificateStatus.EXPIRED:
                status = "expired"
            elif days_left is not None and days_left <= 90:
                status = "expiring"
            else:
                continue
            items.append(
                ExpiryWarningItem(
                    id=cert.id,
                    kind="certificate",
                    title=cert.name,
                    detail="、".join(
                        filter(
                            None,
                            [
                                cert.level,
                                cert.specialty,
                                cert.cert_number and f"编号 {cert.cert_number}",
                            ],
                        )
                    )
                    or cert.issuing_authority
                    or "资质证书",
                    valid_to=cert.valid_to,
                    days_left=days_left,
                    status=status,  # type: ignore[arg-type]
                )
            )

        for item in personnel:
            days_left = (item.valid_to - today).days if item.valid_to else None
            if not item.is_on_job:
                status = "off_job"
            elif days_left is not None and days_left < 0:
                status = "expired"
            elif days_left is not None and days_left <= 90:
                status = "expiring"
            else:
                continue
            items.append(
                ExpiryWarningItem(
                    id=item.id,
                    kind="personnel",
                    title=f"{item.person_name} · {item.cert_type}",
                    detail=item.specialty or item.cert_number or "人员证书",
                    valid_to=item.valid_to,
                    days_left=days_left,
                    status=status,  # type: ignore[arg-type]
                )
            )

        expired_count = sum(
            1 for item in items if item.status in {"expired", "revoked", "off_job"}
        )
        expiring_count = sum(1 for item in items if item.status == "expiring")
        return ExpiryWarningsRead(
            items=items,
            expired_count=expired_count,
            expiring_count=expiring_count,
        )

    @staticmethod
    async def import_excel(
        session: AsyncSession,
        file: UploadFile,
    ) -> QualificationImportResult:
        """按工作表批量导入知识库：资质证书/业绩/人员证书/公司信息。"""

        result = QualificationImportResult()
        content = await file.read()
        try:
            workbook = load_workbook(
                io.BytesIO(content),
                data_only=True,
                read_only=True,
            )
        except Exception as exc:
            result.failed = 1
            result.errors.append(f"无法解析 Excel 文件：{exc}")
            return result

        sheet_specs = [
            (
                ("资质证书", "证书", "certificates"),
                "certificate",
                _CERTIFICATE_HEADERS,
            ),
            (
                ("业绩", "performances"),
                "performance",
                _PERFORMANCE_HEADERS,
            ),
            (
                ("人员证书", "人员", "personnel"),
                "personnel",
                _PERSONNEL_HEADERS,
            ),
            (
                ("公司信息", "公司", "companies"),
                "company",
                _COMPANY_HEADERS,
            ),
        ]
        for sheet_aliases, kind, headers in sheet_specs:
            worksheet = None
            for alias in sheet_aliases:
                if alias in workbook.sheetnames:
                    worksheet = workbook[alias]
                    break
            if worksheet is None:
                continue
            await QualificationService._import_sheet(
                session,
                worksheet,
                kind,
                headers,
                result,
            )

        try:
            await session.commit()
        except Exception as exc:
            await session.rollback()
            result.failed += result.created
            result.created = 0
            result.errors.append(f"批量写入失败：{exc}")
        return result

    @staticmethod
    async def _import_sheet(
        session: AsyncSession,
        worksheet: Any,
        kind: str,
        headers: dict[str, str],
        result: QualificationImportResult,
    ) -> None:
        header_map: dict[int, str] = {}
        data_rows: list[dict[str, Any]] = []
        for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            values = ["" if value is None else value for value in row]
            if not any(str(value).strip() for value in values):
                continue
            if not header_map:
                for column_index, value in enumerate(values):
                    label = str(value).strip()
                    if label in headers:
                        header_map[column_index] = headers[label]
                if not header_map:
                    result.failed += 1
                    result.errors.append(
                        f"{worksheet.title} 首行未找到可识别的表头，跳过该表"
                    )
                    return
                continue
            record_data: dict[str, Any] = {}
            for column_index, value in enumerate(values):
                field = header_map.get(column_index)
                if field:
                    record_data[field] = _normalize_import_value(field, value)
            data_rows.append((row_index, record_data))

        for row_index, record_data in data_rows:
            try:
                await QualificationService._import_row(session, kind, record_data)
                result.created += 1
            except (ValidationError, ValueError, TypeError) as exc:
                result.failed += 1
                result.errors.append(
                    f"{worksheet.title} 第 {row_index} 行导入失败：{_first_error(exc)}"
                )

    @staticmethod
    async def _import_row(
        session: AsyncSession,
        kind: str,
        data: dict[str, Any],
    ) -> None:
        values = {key: value for key, value in data.items() if value is not _OMIT}
        if kind == "certificate":
            payload = QualificationCertificateCreate(**values)
            session.add(QualificationCertificate(**payload.model_dump()))
        elif kind == "performance":
            payload = PerformanceRecordCreate(**values)
            session.add(PerformanceRecord(**payload.model_dump()))
        elif kind == "personnel":
            payload = PersonnelCertificateCreate(**values)
            session.add(PersonnelCertificate(**payload.model_dump()))
        elif kind == "company":
            existing = await session.scalar(
                select(CompanyProfile.id).where(
                    CompanyProfile.company_name == values["company_name"]
                )
            )
            if existing is not None:
                raise ValueError("公司名称已存在")
            payload = CompanyProfileCreate(**values)
            session.add(CompanyProfile(**payload.model_dump()))

    @staticmethod
    async def _paginate(
        session: AsyncSession,
        model: type[ModelT],
        filters: list[Any],
        page: int,
        page_size: int,
        order_by: Any,
    ) -> tuple[list[ModelT], int]:
        total = int(
            (await session.scalar(select(func.count()).select_from(model).where(*filters)))
            or 0
        )
        statement = (
            select(model)
            .where(*filters)
            .order_by(order_by)
            .offset((page - 1) * page_size)
            .limit(page_size)
        )
        return list((await session.scalars(statement)).all()), total

    @staticmethod
    async def _delete_with_attachment(
        session: AsyncSession,
        storage: MinIOStorage,
        record: QualificationCertificate | PerformanceRecord | PersonnelCertificate,
    ) -> None:
        if record.file_object_key:
            try:
                await storage.delete_object(record.file_object_key)
            except Exception as exc:
                await session.rollback()
                raise StorageException("删除资质关联的 MinIO 文件失败") from exc
        await session.delete(record)
        await session.commit()

    @staticmethod
    def _apply_values(record: Any, values: dict[str, Any]) -> None:
        for field, value in values.items():
            setattr(record, field, value)

    @staticmethod
    def _validate_merged_date_range(
        record: Any,
        values: dict[str, Any],
        label: str,
        *,
        start_field: str = "valid_from",
        end_field: str = "valid_to",
    ) -> None:
        start_value = values.get(start_field, getattr(record, start_field))
        end_value = values.get(end_field, getattr(record, end_field))
        if start_value and end_value and start_value > end_value:
            raise AppException(
                f"{label}开始日期不能晚于结束日期",
                code=42221,
                status_code=422,
            )

    @staticmethod
    def _certificate_valid_condition() -> Any:
        today = date.today()
        return and_(
            QualificationCertificate.status == CertificateStatus.VALID,
            or_(
                QualificationCertificate.valid_from.is_(None),
                QualificationCertificate.valid_from <= today,
            ),
            or_(
                QualificationCertificate.valid_to.is_(None),
                QualificationCertificate.valid_to >= today,
            ),
        )

    @staticmethod
    def _personnel_valid_condition() -> Any:
        today = date.today()
        return and_(
            PersonnelCertificate.is_on_job.is_(True),
            or_(
                PersonnelCertificate.valid_from.is_(None),
                PersonnelCertificate.valid_from <= today,
            ),
            or_(
                PersonnelCertificate.valid_to.is_(None),
                PersonnelCertificate.valid_to >= today,
            ),
        )


_OMIT = object()

_CERTIFICATE_HEADERS = {
    "证书名称": "name",
    "名称": "name",
    "等级": "level",
    "级别": "level",
    "专业": "specialty",
    "证书编号": "cert_number",
    "编号": "cert_number",
    "发证机关": "issuing_authority",
    "生效日期": "valid_from",
    "有效期开始": "valid_from",
    "失效日期": "valid_to",
    "有效期至": "valid_to",
    "到期日期": "valid_to",
    "状态": "status",
    "备注": "remark",
}

_PERFORMANCE_HEADERS = {
    "项目名称": "project_name",
    "合同金额": "project_amount",
    "项目金额": "project_amount",
    "金额": "project_amount",
    "币种": "currency",
    "开始日期": "start_date",
    "结束日期": "end_date",
    "是否完成": "is_completed",
    "完成状态": "is_completed",
    "业主名称": "owner_name",
    "业主": "owner_name",
    "采购人": "owner_name",
    "项目地点": "location",
    "地点": "location",
    "关联资质": "related_qualification",
    "项目描述": "description",
    "描述": "description",
}

_PERSONNEL_HEADERS = {
    "姓名": "person_name",
    "人员姓名": "person_name",
    "人员": "person_name",
    "证书类型": "cert_type",
    "证书": "cert_type",
    "专业": "specialty",
    "证书编号": "cert_number",
    "编号": "cert_number",
    "生效日期": "valid_from",
    "有效期开始": "valid_from",
    "失效日期": "valid_to",
    "有效期至": "valid_to",
    "到期日期": "valid_to",
    "是否在职": "is_on_job",
    "在职状态": "is_on_job",
    "备注": "remark",
}

_COMPANY_HEADERS = {
    "公司名称": "company_name",
    "法定代表人": "legal_person",
    "法人": "legal_person",
    "注册资本": "registered_capital",
    "成立日期": "establish_date",
    "地址": "address",
    "联系方式": "contact_info",
    "联系电话": "contact_info",
    "电话": "contact_info",
}


def _normalize_import_value(field: str, value: Any) -> Any:
    defaulted = {"status", "currency", "is_completed", "is_on_job"}
    if value is None:
        return _OMIT if field in defaulted else None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return _OMIT if field in defaulted else None
    if field in {"is_completed", "is_on_job"}:
        return _parse_import_bool(text)
    if field == "status":
        return _parse_import_status(text)
    if field in {"project_amount", "registered_capital"}:
        try:
            return Decimal(text.replace(",", ""))
        except Exception:
            return None
    return text


def _parse_import_bool(value: str) -> bool:
    if value in {"是", "有", "√", "TRUE", "true", "True", "1", "Y", "y"}:
        return True
    if value in {"否", "无", "×", "FALSE", "false", "False", "0", "N", "n"}:
        return False
    raise ValueError(f"无法识别布尔值：{value}")


def _parse_import_status(value: str) -> str:
    mapping = {
        "有效": "valid",
        "正常": "valid",
        "已过期": "expired",
        "过期": "expired",
        "失效": "expired",
        "已撤销": "revoked",
        "撤销": "revoked",
    }
    return mapping.get(value, value)


def _first_error(exc: Exception) -> str:
    if isinstance(exc, ValidationError):
        first = exc.errors()[0]
        location = ".".join(str(part) for part in first.get("loc", ()))
        message = first.get("msg", str(exc))
        return f"{location} {message}".strip()
    return str(exc)
